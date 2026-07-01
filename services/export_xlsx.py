"""Write the finished roster workbook (.xlsx) for pasting into Google Sheets.

Produces three tabs for the target month:
  1. "<Month> <Year> Workday"        — Snowflake/file roster, as VALUES. Single
     landing tab replacing the old manual Workday paste + IMPORTRANGE "Week of".
  2. "<Month> <Year> Playvox Roster" — filtered Playvox, as VALUES. Column layout
     A First / B Last / C Full Name / D Email / E Business Roles, so the roster's
     column L formula (VLOOKUP email in $D:$E) resolves.
  3. "All Teams <Month> <Year>"      — the main roster. Columns F (email) and N
     (comment) are values; everything else is a live formula whose refs point at
     tabs 1, 2, the prior-month roster, and the existing 'Z2 Names List' tab.

The roster header sits on row 6, data starts on row 7 (matching the live sheet),
so a user can paste any tab straight in.
"""

import csv
import datetime
import io

import openpyxl
import pandas as pd

from core import config, refs
from services.playvox import PlayvoxResult


def _clean_csv(value):
    """Coerce a pandas value into a clean CSV cell (resolved value, not formula).

    NA/NaT -> "", dates -> ISO (date-only when midnight), numpy scalars -> native.
    """
    if value is None:
        return ""
    if not isinstance(value, (list, tuple)):
        try:
            if pd.isna(value):
                return ""
        except (TypeError, ValueError):
            pass
    if isinstance(value, pd.Timestamp):
        value = value.to_pydatetime()
    if isinstance(value, datetime.datetime):
        return value.date().isoformat() if value.time() == datetime.time(0, 0) else value.isoformat(sep=" ")
    if isinstance(value, datetime.date):
        return value.isoformat()
    if hasattr(value, "item"):  # numpy scalar -> native python
        try:
            return value.item()
        except (ValueError, AttributeError):
            pass
    # Guard against a literal "nan"/"none" string that slipped in upstream.
    if isinstance(value, str) and value.strip().lower() in ("nan", "none", "nat"):
        return ""
    return value


_MANAGER_INSTRUCTIONS = (
    "Monthly Manager Validation Instructions\n"
    "- Validate all columns for your team are complete and accurate. "
    "Comment in the sheet and tag Giuseppe Eustaquio with any errors\n"
    "- Special attention to Column E (Advocate Z2 name) - update cell directly "
    "if advocate's Z2 name is different from column D\n"
    "- [Core Managers Only] Special attention to Column K (Role) - update cell "
    "directly if role type is different"
)


def _banner_rows(month_num: int, year: int) -> list[list[str]]:
    """The 5 banner rows that sit above the A..N header in the live sheet.

    Header lands on row 6 / data on row 7, matching config.ROSTER_HEADER_ROW.
    Row 3 carries the manager instructions; row 5 the update-date labels
    (Sheet Updated Date pre-filled to the 1st of the month; the Workday "Week of"
    and Explore date are left blank to fill in the sheet).
    """
    sheet_updated = f"{refs.month_name(month_num)} 1, {year}"
    return [
        [],
        [],
        ["", _MANAGER_INSTRUCTIONS],
        [],
        ["", "Workday Source: ", "", "Sheet Updated Date", sheet_updated,
         "Explore Updated Date:", ""],
    ]


def build_csv(roster_df, month_num: int, year: int) -> bytes:
    """Return the resolved roster (A..N, real values) as CSV bytes.

    Holds actual resolved values (Z2 name, Region in Explore, manager, etc.), so
    it opens cleanly in Excel and Google Sheets. Prefixed with the 5 banner rows
    from the live sheet (instructions + update-date labels) so the header lands
    on row 6 / data on row 7 — paste-ready into the roster doc. utf-8-sig so
    Excel detects UTF-8.
    """
    df = roster_df.reindex(columns=list(config.ROSTER_COLUMNS.values())).copy()
    for col in df.columns:
        df[col] = df[col].map(_clean_csv)

    buffer = io.StringIO()
    writer = csv.writer(buffer)
    for row in _banner_rows(month_num, year):
        writer.writerow(row)
    writer.writerow(list(config.ROSTER_COLUMNS.values()))  # header on row 6
    for record in df.itertuples(index=False, name=None):
        writer.writerow(record)
    return buffer.getvalue().encode("utf-8-sig")


def output_csv_filename(month_num: int, year: int) -> str:
    # Matches the Google Sheet tab naming, e.g. "All Teams [July 2026].csv".
    return f"All Teams [{refs.month_name(month_num)} {year}].csv"


def _clean(value):
    """Coerce a pandas value into something openpyxl can write.

    Handles pd.NA / NaT / NaN (which openpyxl rejects) and Timestamps.
    """
    if value is None or (not isinstance(value, (list, tuple)) and pd.isna(value)):
        return ""
    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime()
    if hasattr(value, "item"):  # numpy scalar -> native python
        try:
            return value.item()
        except (ValueError, AttributeError):
            pass
    return value


def _write_workday_tab(wb, tab_name: str, workday_df) -> None:
    ws = wb.create_sheet(tab_name)
    ws.append(list(workday_df.columns))
    for record in workday_df.itertuples(index=False, name=None):
        ws.append([_clean(v) for v in record])


def _write_playvox_tab(wb, tab_name: str, playvox: PlayvoxResult) -> None:
    ws = wb.create_sheet(tab_name)
    # Layout the roster's column-L formula expects: D = email, E = roles.
    ws.append(["First name", "Last name", "Full Name", "Email address", "Business Roles"])
    df = playvox.kept[
        [
            config.PLAYVOX_FIRST_NAME_COL,
            config.PLAYVOX_LAST_NAME_COL,
            "Full Name",
            config.PLAYVOX_EMAIL_COL,
            config.PLAYVOX_ROLE_COL,
        ]
    ]
    for record in df.itertuples(index=False, name=None):
        ws.append([_clean(v) for v in record])


def _write_roster_tab(wb, tab_name: str, roster_df, tabs: dict[str, str]) -> None:
    ws = wb.create_sheet(tab_name)

    # Header on row 6 (rows 1-5 left blank, mirroring the live sheet banner area).
    # openpyxl uses integer columns; map letter -> 1-based index.
    letter_to_idx = {letter: i + 1 for i, letter in enumerate(config.ROSTER_COLUMNS)}
    for letter, header in config.ROSTER_COLUMNS.items():
        ws.cell(row=config.ROSTER_HEADER_ROW, column=letter_to_idx[letter], value=header)

    email_header = config.ROSTER_COLUMNS["F"]
    comment_header = config.ROSTER_COLUMNS["N"]

    for i, record in enumerate(roster_df.to_dict("records")):
        sheet_row = config.ROSTER_DATA_START_ROW + i
        formulas = refs.roster_formulas(sheet_row, tabs)

        for letter, idx in letter_to_idx.items():
            if letter == "F":
                value = _clean(record.get(email_header, ""))
            elif letter == "N":
                value = _clean(record.get(comment_header, ""))
            else:
                value = formulas[letter]
            ws.cell(row=sheet_row, column=idx, value=value)


def build_workbook(
    roster_df,
    workday_df,
    playvox: PlayvoxResult,
    month_num: int,
    year: int,
) -> bytes:
    """Build the workbook and return it as .xlsx bytes (for st.download_button)."""
    tabs = refs.tab_names(month_num, year)

    wb = openpyxl.Workbook()
    # Drop the default empty sheet.
    wb.remove(wb.active)

    _write_roster_tab(wb, tabs["roster"], roster_df, tabs)
    _write_workday_tab(wb, tabs["workday"], workday_df)
    _write_playvox_tab(wb, tabs["playvox"], playvox)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def output_filename(month_num: int, year: int) -> str:
    return f"Roster_{refs.month_name(month_num)}_{year}.xlsx"
