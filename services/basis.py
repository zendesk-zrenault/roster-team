"""Basis file for the two carry-forward roster columns that exist in no upload:
   C "Region in Explore (Shift)" and M "Foreign Language Advocate".

The basis is a prior-roster-shaped .xlsx (roster headers on row 1) keyed by
Advocate Email. Each month the app:
  * carries C and M forward from the basis for known emails,
  * detects roster emails NOT in the basis (new agents),
  * after the user supplies region+language for them, appends those rows to the
    basis so they are known next month.

Fallbacks when an email is absent and no value is supplied: the Workday region
for C, and config.DEFAULT_FOREIGN_LANGUAGE for M.
"""

from pathlib import Path

import openpyxl
import pandas as pd

from core import config

EMAIL_COL = config.ROSTER_COLUMNS["F"]            # "Advocate Email"
REGION_EXPLORE_COL = config.ROSTER_COLUMNS["C"]   # "Region in Explore (Shift)"
LANGUAGE_COL = config.ROSTER_COLUMNS["M"]         # "Foreign Language Advocate"

# Tolerate the header drift seen in the sheet ("Role" vs "Role ").
_BASIS_USECOLS = [EMAIL_COL, REGION_EXPLORE_COL, LANGUAGE_COL]


def basis_exists() -> bool:
    return Path(config.BASIS_FILE).exists()


def load_basis() -> pd.DataFrame:
    """Return the basis as a DataFrame [Email, Region in Explore, Language].

    Emails are lowercased/stripped. Empty frame if the file is missing.
    """
    if not basis_exists():
        return pd.DataFrame(columns=_BASIS_USECOLS)

    df = pd.read_excel(config.BASIS_FILE, header=config.BASIS_HEADER_ROW - 1, dtype=object)
    df.columns = [str(c).strip() for c in df.columns]

    # Be forgiving about which of the three columns are present.
    for col in _BASIS_USECOLS:
        if col not in df.columns:
            df[col] = pd.NA
    df = df[_BASIS_USECOLS].copy()

    df[EMAIL_COL] = df[EMAIL_COL].astype("string").str.strip().str.lower()
    df = df[df[EMAIL_COL].notna() & (df[EMAIL_COL] != "")]
    return df.drop_duplicates(subset=EMAIL_COL, keep="last").reset_index(drop=True)


def region_map() -> dict[str, str]:
    """{email: Region in Explore} from the basis (blanks dropped)."""
    df = load_basis()
    out = {}
    for email, region in zip(df[EMAIL_COL], df[REGION_EXPLORE_COL], strict=False):
        if email and pd.notna(region) and str(region).strip():
            out[email] = str(region).strip()
    return out


def language_map() -> dict[str, str]:
    """{email: Foreign Language} from the basis (blanks dropped)."""
    df = load_basis()
    out = {}
    for email, lang in zip(df[EMAIL_COL], df[LANGUAGE_COL], strict=False):
        if email and pd.notna(lang) and str(lang).strip():
            out[email] = str(lang).strip()
    return out


def known_emails() -> set[str]:
    return set(load_basis()[EMAIL_COL])


def existing_regions() -> list[str]:
    """Distinct region values already in the basis (for dropdown options)."""
    return sorted({v for v in region_map().values()})


def existing_languages() -> list[str]:
    """Distinct language values already in the basis (for dropdown options)."""
    return sorted({v for v in language_map().values()})


def find_new_agents(
    roster: pd.DataFrame, eligible_emails: set[str] | None = None
) -> pd.DataFrame:
    """Roster rows whose email is NOT in the basis — these need region+language.

    Only agents whose email is in `eligible_emails` are returned (pass the
    Playvox/Agyle kept emails, i.e. agents present there with a valid business
    role). Workday-only agents are not prompted; they just take the fallbacks.

    Returns [Advocate Email, Advocate, Region in Workday] — the Workday region is
    the suggested default for Region in Explore.
    """
    have = known_emails()
    mask = ~roster[EMAIL_COL].isin(have)
    if eligible_emails is not None:
        eligible = {e.strip().lower() for e in eligible_emails if e}
        mask &= roster[EMAIL_COL].isin(eligible)
    cols = [EMAIL_COL, config.ROSTER_COLUMNS["D"], config.ROSTER_COLUMNS["B"]]
    return roster.loc[mask, cols].reset_index(drop=True)


def append_entries(entries: list[dict]) -> int:
    """Append new agent rows to the basis workbook and save in place.

    `entries` items: {"email": str, "region": str, "language": str}. Existing
    emails are skipped. Returns the number of rows added. Writes a roster-shaped
    sheet so the file stays reusable as a prior roster.
    """
    new = [
        e for e in entries
        if e.get("email") and e["email"].strip().lower() not in known_emails()
    ]
    if not new:
        return 0

    path = Path(config.BASIS_FILE)
    if path.exists():
        wb = openpyxl.load_workbook(path)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(list(config.ROSTER_COLUMNS.values()))

    # Map header label -> 1-based column index from the basis header row.
    headers = [
        (ws.cell(row=config.BASIS_HEADER_ROW, column=c).value or "").strip()
        for c in range(1, ws.max_column + 1)
    ]
    idx = {h: i + 1 for i, h in enumerate(headers)}

    for e in new:
        row = ws.max_row + 1
        if EMAIL_COL in idx:
            ws.cell(row=row, column=idx[EMAIL_COL], value=e["email"].strip().lower())
        if REGION_EXPLORE_COL in idx:
            ws.cell(row=row, column=idx[REGION_EXPLORE_COL], value=e.get("region", ""))
        if LANGUAGE_COL in idx:
            ws.cell(row=row, column=idx[LANGUAGE_COL], value=e.get("language", ""))

    wb.save(path)
    return len(new)
