"""Seed data/z2_names_list.csv from a roster workbook's 'Z2 Names List' tab.

The Z2 Names List maps Advocate Email -> Zendesk/Z2 display name (column E in the
roster). It is the fallback when no Snowflake Zendesk-users lookup is available,
and the app appends newly-resolved names back into it each month.

Usage:
    python scripts/seed_z2_cache.py "samples/Claude Copy of Explore Roster Teams.xlsx"
"""

import sys
from pathlib import Path

import openpyxl
import pandas as pd

TAB = "Z2 Names List"
OUT = Path("data/z2_names_list.csv")


def main(workbook_path: str) -> None:
    wb = openpyxl.load_workbook(workbook_path, data_only=True, read_only=True)
    if TAB not in wb.sheetnames:
        raise SystemExit(f"Tab {TAB!r} not found in {workbook_path}")
    ws = wb[TAB]

    records = []
    for email, name in ws.iter_rows(min_row=2, max_col=2, values_only=True):
        if email and name:
            records.append({"Email": str(email).strip().lower(), "Z2 Name": str(name).strip()})
    wb.close()

    df = pd.DataFrame(records, columns=["Email", "Z2 Name"])
    # Keep the last occurrence per email (most recent wins on duplicates).
    df = df.drop_duplicates(subset="Email", keep="last").reset_index(drop=True)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(OUT, index=False)
    print(f"Wrote {len(df):,} unique email->name pairs to {OUT}")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        raise SystemExit("Usage: python scripts/seed_z2_cache.py <roster_workbook.xlsx>")
    main(sys.argv[1])
