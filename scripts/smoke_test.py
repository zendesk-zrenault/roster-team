"""Headless end-to-end smoke test of the full roster pipeline.

Exercises the service layer exactly as the app does, using the sample files and
the XLS-fallback Workday path (Snowflake needs live auth). Asserts the key
invariants verified during planning. Run:

    .venv/bin/python scripts/smoke_test.py
"""

import io
import sys
import warnings
from pathlib import Path

warnings.filterwarnings("ignore")
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl  # noqa: E402
import pandas as pd  # noqa: E402

from services.export_xlsx import build_csv, build_workbook  # noqa: E402
from services.playvox import load_playvox_csv  # noqa: E402
from services.roster_builder import build_roster, find_unresolved, split_departed  # noqa: E402
from services.workday_file import load_workday_file  # noqa: E402

WORKDAY = "samples/Global Advocacy Headcount (Madison) 2026-06-18 02_01 CST.xlsx"
PLAYVOX = "samples/playvox_users_export_2026-06-19T09_53_49.csv"


def main() -> None:
    wd = load_workday_file(WORKDAY)
    assert len(wd) == 404, f"expected 404 advocacy rows, got {len(wd)}"
    cc = wd["COST_CENTER"].value_counts()
    assert cc.get("280 Core") == 258, cc.to_dict()
    assert cc.get("287 Premier") == 101, cc.to_dict()
    print(f"✓ Workday XLS fallback: {len(wd)} advocacy rows (280 Core=258, 287 Premier=101)")

    pv = load_playvox_csv(PLAYVOX)
    assert pv.kept_count == 334, f"expected 334 kept, got {pv.kept_count}"
    print(f"✓ Playvox filter: kept {pv.kept_count}/{pv.total_rows}")

    roster = build_roster(wd, pv, z2_map={})  # z2 cache is Snowflake-backed; skip in offline test
    email_col = "Advocate Email"
    assert roster[email_col].is_unique, "duplicate emails survived dedup"
    print(f"✓ Roster built: {len(roster)} unique rows")

    # Job-title filter: drop non-ticket-bearing / management rows (CSV fallback).
    from services import job_titles  # noqa: PLC0415
    tb, ntb = job_titles.split_by_ticket_bearing(roster)
    unknown = job_titles.unclassified_titles(roster)
    assert len(tb) + len(ntb) == len(roster), (len(tb), len(ntb), len(roster))
    print(f"✓ Job-title filter: {len(tb)} ticket-bearing kept, {len(ntb)} dropped, "
          f"{len(unknown)} unclassified")

    present, departed = split_departed(tb, set(wd["EMAIL"].dropna()))
    unresolved = find_unresolved(tb)
    print(f"✓ Review split: present={len(present)} departed={len(departed)} unresolved={len(unresolved)}")

    data = build_workbook(roster, wd, pv, month_num=7, year=2026)
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=False)
    assert wb.sheetnames == [
        "All Teams July 2026",
        "July 2026 Workday",
        "July 2026 Playvox Roster",
    ], wb.sheetnames
    ws = wb["All Teams July 2026"]
    # Formula sanity: B references the July Workday tab; L references the July Playvox tab.
    assert "'July 2026 Workday'" in ws["B7"].value, ws["B7"].value
    assert "'July 2026 Playvox Roster'" in ws["L7"].value, ws["L7"].value
    assert "'All Teams June 2026'" in ws["M7"].value, ws["M7"].value  # prior-month carry-forward
    wb.close()
    print(f"✓ Export (.xlsx): 3 tabs, formulas reference correct month tabs ({len(data):,} bytes)")

    # CSV export: 5 banner rows, header on row 6, resolved values, no formulas.
    csv_bytes = build_csv(roster, month_num=7, year=2026)
    # header row is index 5 (0-based) after the 5 banner rows.
    csv_df = pd.read_csv(io.BytesIO(csv_bytes), header=5, skip_blank_lines=False)
    assert list(csv_df.columns) == list(roster.columns), csv_df.columns.tolist()
    assert len(csv_df) == len(roster), (len(csv_df), len(roster))
    assert not csv_df.astype(str).apply(lambda c: c.str.startswith("=")).any().any(), (
        "CSV must contain resolved values, not formulas"
    )
    # No literal "nan" strings leaked into any cell.
    assert not csv_df.astype(str).apply(lambda c: c.str.strip().str.lower() == "nan").any().any(), (
        "CSV must not contain literal 'nan' strings"
    )
    print(f"✓ Export (.csv): 5 banner rows + header on row 6, {len(csv_df)} rows, "
          f"no formulas/nan ({len(csv_bytes):,} bytes)")

    print("\nALL CHECKS PASSED")


if __name__ == "__main__":
    main()
